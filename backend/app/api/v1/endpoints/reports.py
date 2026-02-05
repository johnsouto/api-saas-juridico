from __future__ import annotations

import io
from typing import Annotated

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.agenda_evento import AgendaEvento
from app.models.client import Client
from app.models.honorario import Honorario
from app.models.parceria import Parceria
from app.models.process import Process
from app.models.tarefa import Tarefa
from app.models.user import User


router = APIRouter()


@router.get("/overview.xlsx")
async def export_overview_xlsx(
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(get_current_user)],
):
    """
    Export a basic office overview as an .xlsx file (tenant-scoped).
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Clientes
    ws = wb.create_sheet("Clientes")
    ws.append(["id", "nome", "cpf", "criado_em"])
    clients = list((await db.execute(select(Client).where(Client.tenant_id == user.tenant_id).order_by(Client.criado_em.desc()))).scalars().all())
    for c in clients:
        ws.append([str(c.id), c.nome, c.cpf, c.criado_em.isoformat()])

    # Parcerias
    ws = wb.create_sheet("Parcerias")
    ws.append(["id", "nome", "email", "telefone", "tipo_documento", "documento", "criado_em"])
    partners = list((await db.execute(select(Parceria).where(Parceria.tenant_id == user.tenant_id).order_by(Parceria.criado_em.desc()))).scalars().all())
    for p in partners:
        ws.append([str(p.id), p.nome, p.email, p.telefone, p.tipo_documento.value, p.documento, p.criado_em.isoformat()])

    # Processos
    ws = wb.create_sheet("Processos")
    ws.append(["id", "numero", "status", "client_id", "parceria_id", "criado_em"])
    procs = list((await db.execute(select(Process).where(Process.tenant_id == user.tenant_id).order_by(Process.criado_em.desc()))).scalars().all())
    for pr in procs:
        ws.append([str(pr.id), pr.numero, pr.status, str(pr.client_id), str(pr.parceria_id) if pr.parceria_id else "", pr.criado_em.isoformat()])

    # Honorários
    ws = wb.create_sheet("Honorarios")
    ws.append(
        [
            "id",
            "process_id",
            "valor_inicial",
            "qtd_parcelas",
            "data_inicio_pagamento",
            "status",
            "percentual_exito",
            "percentual_parceiro",
            "pago_em",
            "valor_pago",
            "meio_pagamento",
            "criado_em",
        ]
    )
    honorarios = list((await db.execute(select(Honorario).where(Honorario.tenant_id == user.tenant_id).order_by(Honorario.criado_em.desc()))).scalars().all())
    for h in honorarios:
        ws.append(
            [
                str(h.id),
                str(h.process_id),
                str(h.valor),
                int(h.qtd_parcelas),
                h.data_vencimento.isoformat(),
                h.status.value,
                h.percentual_exito if h.percentual_exito is not None else "",
                h.percentual_parceiro if h.percentual_parceiro is not None else "",
                h.pago_em.isoformat() if h.pago_em else "",
                str(h.valor_pago) if h.valor_pago is not None else "",
                h.meio_pagamento or "",
                h.criado_em.isoformat(),
            ]
        )

    # Agenda
    ws = wb.create_sheet("Agenda")
    ws.append(["id", "titulo", "tipo", "inicio_em", "fim_em", "process_id", "client_id", "descricao"])
    eventos = list((await db.execute(select(AgendaEvento).where(AgendaEvento.tenant_id == user.tenant_id).order_by(AgendaEvento.inicio_em.asc()))).scalars().all())
    for e in eventos:
        ws.append(
            [
                str(e.id),
                e.titulo,
                e.tipo,
                e.inicio_em.isoformat(),
                e.fim_em.isoformat() if e.fim_em else "",
                str(e.process_id) if e.process_id else "",
                str(e.client_id) if e.client_id else "",
                e.descricao or "",
            ]
        )

    # Tarefas
    ws = wb.create_sheet("Tarefas")
    ws.append(["id", "titulo", "status", "prazo_em", "client_id", "responsavel_id", "criado_em"])
    tarefas = list((await db.execute(select(Tarefa).where(Tarefa.tenant_id == user.tenant_id).order_by(Tarefa.criado_em.desc()))).scalars().all())
    for t in tarefas:
        ws.append(
            [
                str(t.id),
                t.titulo,
                t.status.value,
                t.prazo_em.isoformat() if t.prazo_em else "",
                str(t.client_id) if t.client_id else "",
                str(t.responsavel_id) if t.responsavel_id else "",
                t.criado_em.isoformat(),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers = {"Content-Disposition": 'attachment; filename="relatorio.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

